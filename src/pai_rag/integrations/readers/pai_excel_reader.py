"""Tabular parser-Excel parser.

Contains parsers for tabular data files.

"""

from pathlib import Path
from typing import Any, Dict, List, Optional
from fsspec import AbstractFileSystem
from openpyxl import load_workbook

import pandas as pd
from llama_index.core.readers.base import BaseReader
from llama_index.core.schema import Document


class PaiPandasExcelReader(BaseReader):
    r"""Pandas-based Excel parser.


    Args:
        concat_rows (bool): whether to concatenate all rows into one document.
            If set to False, a Document will be created for each row.
            True by default.

        row_joiner (str): Separator to use for joining each row.
            Only used when `concat_rows=True`.
            Set to "\n" by default.

        pandas_config (dict): Options for the `pandas.read_excel` function call.
            Refer to https://pandas.pydata.org/docs/reference/api/pandas.read_excel.html
            for more information.
            Set to empty dict by default, this means pandas will try to figure
            out the separators, table head, etc. on its own.

    """

    def __init__(
        self,
        *args: Any,
        concat_rows: bool = True,
        row_joiner: str = "\n",
        pandas_config: dict = {},
        format_sheet_data_to_json: bool = False,
        sheet_column_filters: List[str] = None,
        **kwargs: Any,
    ) -> None:
        """Init params."""
        super().__init__(*args, **kwargs)
        self._concat_rows = concat_rows
        self._row_joiner = row_joiner
        self._pandas_config = pandas_config
        self._format_sheet_data_to_json = format_sheet_data_to_json
        self._sheet_column_filters = sheet_column_filters

    def read_xlsx(
        self,
        file: Path,
        fs: Optional[AbstractFileSystem] = None,
    ):
        """Parse Excel file。"""
        if fs:
            with fs.open(file) as f:
                excel = pd.ExcelFile(load_workbook(f), engine="openpyxl")
        else:
            excel = pd.ExcelFile(load_workbook(file), engine="openpyxl")
        sheet_name = excel.sheet_names[0]
        sheet = excel.book[sheet_name]
        df = excel.parse(sheet_name, **self._pandas_config)

        header_max = 0
        if (
            "header" in self._pandas_config
            and self._pandas_config["header"] is not None
            and isinstance(self._pandas_config["header"], list)
        ):
            header_max = max(self._pandas_config["header"])
        elif (
            "header" in self._pandas_config
            and self._pandas_config["header"] is not None
            and isinstance(self._pandas_config["header"], int)
        ):
            header_max = self._pandas_config["header"]

        for item in sheet.merged_cells:
            top_col, top_row, bottom_col, bottom_row = item.bounds
            base_value = item.start_cell.value
            # Convert 1-based index to 0-based index
            top_row -= 1
            top_col -= 1
            # Since the previous lines are set as headers, the coordinates need to be adjusted here.
            if (
                "header" in self._pandas_config
                and self._pandas_config["header"] is not None
            ) or "header" not in self._pandas_config:
                top_row -= header_max + 1
                bottom_row -= header_max + 1

            df.iloc[top_row:bottom_row, top_col:bottom_col] = base_value
        return df

    def load_data(
        self,
        file: Path,
        extra_info: Optional[Dict] = None,
        fs: Optional[AbstractFileSystem] = None,
    ) -> List[Document]:
        """Parse Excel file. only process the first sheet"""

        df = self.read_xlsx(file, fs)

        if self._sheet_column_filters:
            df = df[self._sheet_column_filters]

        if self._format_sheet_data_to_json:
            text_list = df.apply(
                lambda row: str(dict(zip(df.columns, row.astype(str)))), axis=1
            ).tolist()
        else:
            text_list = [
                "\n".join([f"{k}:{v}" for k, v in record.items()])
                for record in df.to_dict("records")
            ]

        if self._concat_rows:
            return [
                Document(
                    text=(self._row_joiner).join(text_list), metadata=extra_info or {}
                )
            ]
        else:
            docs = []
            for i, text in enumerate(text_list):
                extra_info["row_number"] = i + 1
                docs.append(Document(text=text, metadata=extra_info))
            return docs
